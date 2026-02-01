from __future__ import annotations

from pathlib import Path
from typing import Dict, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# These should match the shift IDs used by your solver
GONDOLA_SHIFTS = {"AM1", "AM2", "MC1_GON", "MC2", "PM1", "PM2"}
GS_SHIFTS = {"TILL1", "TILL2", "TILL3", "GATE", "FLOOR", "FLOOR2", "MC1_GS"}


def _write_value_safe(ws, row: int, col: int, value: str) -> None:
    """Write to a cell, but if it's part of a merged range, write to the top-left anchor."""
    cell = ws.cell(row, col)

    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                ws.cell(merged.min_row, merged.min_col).value = value
                return
        return

    cell.value = value


def _build_name_to_row(ws, name_col: int = 1, start_row: int = 5, max_scan: int = 300) -> dict[str, int]:
    """Read names from column A starting at row 5, return name -> row index."""
    m: dict[str, int] = {}
    for r in range(start_row, start_row + max_scan):
        v = ws.cell(r, name_col).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = r
    return m


def _build_day_to_col(ws, header_row: int = 4, start_col: int = 3, max_scan: int = 30) -> dict[str, int]:
    """Read day headers from row 4, return day -> column index."""
    m: dict[str, int] = {}
    for c in range(start_col, start_col + max_scan):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        day = str(v).strip()
        if day:
            m[day] = c
    return m


def export_roster_to_template(
    assignments: Dict[Tuple[str, str], Optional[str]],
    template_path: str | Path,
    output_path: str | Path,
    gondola_sheet: str = "Gondola",
    gs_sheet: str = "Guest Services",
    gondola_gs_label: str = "GS Host",
) -> None:
    """Export (day, shift)->person assignments to the Excel template.

    - Gondola sheet: writes gondola shifts; if a gondola-visible person is allocated a GS shift,
      writes `gondola_gs_label`.
    - GS sheet: writes GS shift codes for GS staff.
    - If a slot is unfilled (person is None), the cell is left blank.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    wb = load_workbook(template_path)

    ws_g = wb[gondola_sheet]
    ws_s = wb[gs_sheet]

    g_name_to_row = _build_name_to_row(ws_g)
    s_name_to_row = _build_name_to_row(ws_s)

    g_day_to_col = _build_day_to_col(ws_g)
    s_day_to_col = _build_day_to_col(ws_s)

    def clear_sheet(ws, name_to_row, day_to_col):
        for _, r in name_to_row.items():
            for _, c in day_to_col.items():
                _write_value_safe(ws, r, c, "")

    clear_sheet(ws_g, g_name_to_row, g_day_to_col)
    clear_sheet(ws_s, s_name_to_row, s_day_to_col)

    for (day, shift), person in assignments.items():
        if not person:
            continue

        # Gondola sheet
        if person in g_name_to_row and day in g_day_to_col:
            r = g_name_to_row[person]
            c = g_day_to_col[day]
            if shift in GONDOLA_SHIFTS:
                _write_value_safe(ws_g, r, c, shift)
            elif shift in GS_SHIFTS:
                _write_value_safe(ws_g, r, c, gondola_gs_label)
            else:
                _write_value_safe(ws_g, r, c, shift)

        # GS sheet
        if person in s_name_to_row and day in s_day_to_col:
            r = s_name_to_row[person]
            c = s_day_to_col[day]
            if shift in GS_SHIFTS:
                _write_value_safe(ws_s, r, c, shift)

    wb.save(output_path)
