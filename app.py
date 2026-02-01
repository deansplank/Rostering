import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
import shutil
from copy import copy

from openpyxl import load_workbook

from generate_roster import (
    solve_week,
    load_staff,
    load_rules,
    load_requests,
    load_week_template,
    DAYS,
    GONDOLA_SHIFTS,
    GS_SHIFTS,
)
from export_excel import export_roster_to_template


# ============================================================
# CONFIG
# ============================================================
BASE = Path('.')
STAFF_PATH = BASE / 'staff.csv'
RULES_PATH = BASE / 'rules.csv'
WEEK_PATH = BASE / 'week_template.json'
AVAIL_PATH = BASE / 'availability.csv'
REQ_PATH = BASE / 'requests.csv'
TEMPLATE_PATH = BASE / 'roster_template_FINAL.xlsx'

SHIFT_OPTIONS = [
    'ANY',
    'AM1', 'AM2', 'MC1_GON', 'MC2', 'PM1', 'PM2',
    'TILL1', 'TILL2', 'TILL3', 'GATE', 'FLOOR', 'FLOOR2', 'MC1_GS',
]
TYPE_OPTIONS = ['OFF', 'WANT', 'AVOID']


# ============================================================
# Helpers
# ============================================================

def output_filename() -> Path:
    ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    return BASE / f'Generated_Roster_{ts}.xlsx'


def ensure_staff_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure staff.csv supports capability + visibility flags."""
    if 'can_gondola' not in df.columns:
        df['can_gondola'] = 0
    if 'can_gs' not in df.columns:
        df['can_gs'] = 1
    if 'show_gondola' not in df.columns:
        df['show_gondola'] = df['can_gondola']
    if 'show_gs' not in df.columns:
        df['show_gs'] = df['can_gs']

    df['name'] = df['name'].astype(str).str.strip()
    df = df[df['name'] != ''].reset_index(drop=True)

    df['can_gondola'] = df['can_gondola'].fillna(0).astype(int)
    df['can_gs'] = df['can_gs'].fillna(1).astype(int)
    df['show_gondola'] = df['show_gondola'].fillna(df['can_gondola']).astype(int)
    df['show_gs'] = df['show_gs'].fillna(df['can_gs']).astype(int)
    return df


def sync_availability(staff_df: pd.DataFrame) -> pd.DataFrame:
    """Keep availability.csv aligned with staff.csv (Tick = available)."""
    names = staff_df['name'].tolist()

    if AVAIL_PATH.exists():
        avail = pd.read_csv(AVAIL_PATH)
    else:
        avail = pd.DataFrame(columns=['name'] + DAYS)

    if 'name' not in avail.columns:
        avail.insert(0, 'name', '')

    for d in DAYS:
        if d not in avail.columns:
            avail[d] = 1

    avail['name'] = avail['name'].astype(str).str.strip()
    avail = avail[avail['name'].isin(names)]

    existing = set(avail['name'].tolist())
    missing = [n for n in names if n not in existing]
    if missing:
        add_rows = pd.DataFrame([{'name': n, **{d: 1 for d in DAYS}} for n in missing])
        avail = pd.concat([avail, add_rows], ignore_index=True)

    # Keep same order as staff.csv
    order_map = {n: i for i, n in enumerate(names)}
    avail['__order'] = avail['name'].map(lambda x: order_map.get(x, 9999))
    avail = avail.sort_values('__order').drop(columns='__order').reset_index(drop=True)

    avail.to_csv(AVAIL_PATH, index=False)
    return avail


def detect_sheets(template: Path) -> tuple[str, str]:
    wb = load_workbook(template)
    sheets = wb.sheetnames
    gondola = next((s for s in sheets if 'gondola' in s.lower()), sheets[0])
    gs = next(
        (s for s in sheets if 'guest' in s.lower() or s.lower().startswith('gs')),
        sheets[1] if len(sheets) > 1 else sheets[0],
    )
    return gondola, gs


def make_runtime_template() -> Path:
    rt = BASE / '_template_runtime.xlsx'
    shutil.copyfile(TEMPLATE_PATH, rt)
    return rt


def _sheet_day_cols(ws) -> dict[str, int]:
    day_to_col: dict[str, int] = {}
    for c in range(1, 80):
        v = ws.cell(4, c).value
        if v is None:
            continue
        s = str(v).strip()
        if s in DAYS:
            day_to_col[s] = c
    return day_to_col


def _sheet_name_rows(ws, start_row: int = 5, name_col: int = 1, max_scan: int = 900) -> dict[str, int]:
    name_to_row: dict[str, int] = {}
    for r in range(start_row, start_row + max_scan):
        v = ws.cell(r, name_col).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            name_to_row[name] = r
    return name_to_row


def _find_last_staff_row(ws, start_row: int = 5, name_col: int = 1, max_scan: int = 900) -> int:
    last = start_row - 1
    for r in range(start_row, start_row + max_scan):
        v = ws.cell(r, name_col).value
        if v is not None and str(v).strip():
            last = r
    return max(last, start_row)


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    """Copy cell styles from src_row -> dst_row without triggering StyleProxy issues."""
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst._style = copy(src._style)


def apply_visibility_to_runtime_template(runtime_template: Path, gondola_sheet: str, gs_sheet: str, staff_df: pd.DataFrame) -> None:
    """Apply show_gondola/show_gs visibility to the runtime template copy.

    - Clear hidden staff from each sheet (name + day cells)
    - Append missing visible staff to bottom, copying formatting

    This only touches the runtime copy (never the master template).
    """
    staff_df = ensure_staff_columns(staff_df.copy())

    show_gondola = [n for n in staff_df.loc[staff_df['show_gondola'] == 1, 'name'].tolist() if n]
    show_gs = [n for n in staff_df.loc[staff_df['show_gs'] == 1, 'name'].tolist() if n]

    wb = load_workbook(runtime_template)
    ws_g = wb[gondola_sheet]
    ws_s = wb[gs_sheet]

    def process_sheet(ws, allowed_names: list[str]) -> None:
        allowed_set = set(allowed_names)
        day_cols = _sheet_day_cols(ws)
        name_rows = _sheet_name_rows(ws)

        # Clear any names not allowed on this sheet
        for name, r in list(name_rows.items()):
            if name not in allowed_set:
                ws.cell(r, 1).value = ''
                for c in day_cols.values():
                    ws.cell(r, c).value = ''

        # Re-scan
        name_rows = _sheet_name_rows(ws)

        last_staff_row = _find_last_staff_row(ws)
        template_row = last_staff_row
        max_col = max(day_cols.values()) if day_cols else 3 + len(DAYS) - 1

        for name in allowed_names:
            if name in name_rows:
                continue

            new_row = last_staff_row + 1
            last_staff_row = new_row

            _copy_row_style(ws, template_row, new_row, max_col)
            ws.cell(new_row, 1).value = name
            for c in day_cols.values():
                ws.cell(new_row, c).value = ''

    process_sheet(ws_g, show_gondola)
    process_sheet(ws_s, show_gs)
    wb.save(runtime_template)


def build_shift_order(week: dict[str, list[str]]) -> list[str]:
    """Return a stable shift order based on the template's day lists."""
    ordered: list[str] = []
    seen = set()
    for d in DAYS:
        for s in week.get(d, []):
            if s not in seen:
                ordered.append(s)
                seen.add(s)
    return ordered


def assignments_to_table(assignments: dict[tuple[str, str], str | None], week: dict[str, list[str]]) -> pd.DataFrame:
    shift_order = build_shift_order(week)
    df = pd.DataFrame(index=shift_order, columns=DAYS)
    for d in DAYS:
        for s in shift_order:
            if s not in week.get(d, []):
                df.loc[s, d] = ''
            else:
                v = assignments.get((d, s))
                df.loc[s, d] = v if v is not None else '—'
    return df


def people_not_working_that_day(assignments: dict[tuple[str, str], str | None], day: str, people_names: list[str]) -> list[str]:
    working = {p for (d, _s), p in assignments.items() if d == day and p}
    return [p for p in people_names if p not in working]


def published_table_for_sheet(
    staff_df: pd.DataFrame,
    assignments: dict[tuple[str, str], str | None],
    week: dict[str, list[str]],
    show_col: str,
    shift_set: set[str],
) -> pd.DataFrame:
    """Names as rows, days as cols (like Excel)."""
    staff_df = staff_df.copy()
    names = staff_df.loc[staff_df[show_col] == 1, 'name'].tolist()
    rows = []
    for nm in names:
        row = {'Name': nm}
        for d in DAYS:
            val = ''
            for s in week.get(d, []):
                if assignments.get((d, s)) == nm and s in shift_set:
                    val = s
                    break
            row[d] = val
        rows.append(row)
    return pd.DataFrame(rows)


def style_holes(df: pd.DataFrame):
    def _cell(v):
        if v == '—':
            return 'background-color: rgba(255, 0, 0, 0.25);'
        return ''
    return df.style.applymap(_cell)


# ============================================================
# Streamlit
# ============================================================
st.set_page_config(page_title='Skyline Roster Generator', layout='wide')
st.title('🗓️ Skyline Roster Generator')

# Session state
if 'roster' not in st.session_state:
    st.session_state.roster = None  # dict[(day, shift)] -> name|None
if 'seed' not in st.session_state:
    st.session_state.seed = 42

# Load core data
staff_df = pd.read_csv(STAFF_PATH)
staff_df = ensure_staff_columns(staff_df)
staff_df.to_csv(STAFF_PATH, index=False)

avail_df = sync_availability(staff_df)

# ============
# Advanced management (rare)
# ============
with st.expander('Advanced: staff, availability, requests', expanded=False):
    st.subheader('Manage staff')

    col_a, col_b = st.columns(2)
    with col_a:
        with st.form('add_staff', clear_on_submit=True):
            name = st.text_input('Full name').strip()
            can_gondola = st.checkbox('Can work Gondola', value=False)
            can_gs = st.checkbox('Can work Guest Services', value=True)
            show_gondola = st.checkbox('Show on Gondola roster', value=can_gondola)
            show_gs = st.checkbox('Show on Guest Services roster', value=can_gs)
            submit = st.form_submit_button('Add staff')

        if submit:
            if not name:
                st.error('Name required.')
            elif name.lower() in staff_df['name'].str.lower().tolist():
                st.error('Staff member already exists.')
            else:
                staff_df.loc[len(staff_df)] = {
                    'name': name,
                    'can_gondola': int(can_gondola),
                    'can_gs': int(can_gs),
                    'show_gondola': int(show_gondola),
                    'show_gs': int(show_gs),
                }
                staff_df = ensure_staff_columns(staff_df)
                staff_df.to_csv(STAFF_PATH, index=False)
                sync_availability(staff_df)
                st.success(f'Added {name}')
                st.rerun()

    with col_b:
        remove = st.selectbox('Remove staff', [''] + staff_df['name'].tolist())
        confirm = st.checkbox('Confirm removal')
        if st.button('Remove selected'):
            if not remove or not confirm:
                st.error('Select a staff member and confirm.')
            else:
                staff_df = staff_df[staff_df['name'] != remove].reset_index(drop=True)
                staff_df.to_csv(STAFF_PATH, index=False)
                sync_availability(staff_df)
                st.success(f'Removed {remove}')
                st.rerun()

    st.subheader('Edit staff settings (capability + visibility)')
    staff_edit = staff_df.copy()
    staff_edit = st.data_editor(
        staff_edit,
        hide_index=True,
        use_container_width=True,
        column_config={
            'name': st.column_config.TextColumn('Name', disabled=True),
            'can_gondola': st.column_config.CheckboxColumn('can_gondola'),
            'can_gs': st.column_config.CheckboxColumn('can_gs'),
            'show_gondola': st.column_config.CheckboxColumn('show_gondola'),
            'show_gs': st.column_config.CheckboxColumn('show_gs'),
        },
        key='staff_settings_editor',
    )
    staff_edit = ensure_staff_columns(staff_edit)

    if st.button('Save staff settings'):
        staff_edit.to_csv(STAFF_PATH, index=False)
        sync_availability(staff_edit)
        st.success('Staff settings saved')
        st.rerun()

    st.subheader('Availability (tick = available)')
    avail_edit = avail_df.copy()
    for d in DAYS:
        avail_edit[d] = avail_edit[d].fillna(1).astype(int).astype(bool)

    edited_avail = st.data_editor(
        avail_edit,
        disabled=['name'],
        use_container_width=True,
        column_config={d: st.column_config.CheckboxColumn(d) for d in DAYS},
        key='availability_editor',
    )
    if st.button('Save availability'):
        save = edited_avail.copy()
        for d in DAYS:
            save[d] = save[d].fillna(True).astype(bool).astype(int)
        save.to_csv(AVAIL_PATH, index=False)
        st.success('Availability saved')

    st.subheader('Requests')
    if REQ_PATH.exists():
        req_df = pd.read_csv(REQ_PATH)
    else:
        req_df = pd.DataFrame(columns=['name', 'day', 'type', 'shift', 'weight'])

    req_df = st.data_editor(
        req_df,
        num_rows='dynamic',
        use_container_width=True,
        column_config={
            'name': st.column_config.SelectboxColumn('name', options=staff_df['name'].tolist()),
            'day': st.column_config.SelectboxColumn('day', options=DAYS),
            'type': st.column_config.SelectboxColumn('type', options=TYPE_OPTIONS),
            'shift': st.column_config.SelectboxColumn('shift', options=SHIFT_OPTIONS),
            'weight': st.column_config.NumberColumn('weight', min_value=0, step=1),
        },
        key='requests_editor',
    )

    if st.button('Save requests'):
        req_df.to_csv(REQ_PATH, index=False)
        st.success('Requests saved')


# ============================================================
# Generate / Re-roll
# ============================================================
st.header('Generate')

col1, col2, col3 = st.columns([2, 2, 3])
with col1:
    st.session_state.seed = st.slider('Randomness seed', 0, 1000, int(st.session_state.seed))
with col2:
    generate = st.button('🚀 Generate', use_container_width=True)
with col3:
    reroll = st.button('🎲 Re-roll (new holes)', use_container_width=True, disabled=st.session_state.roster is None)


def _build_combined_requests(tmp_path: Path, avail_bool_df: pd.DataFrame, req_df_in: pd.DataFrame) -> None:
    # Availability -> OFF requests
    off_rows = []
    for _, r in avail_bool_df.iterrows():
        for d in DAYS:
            if not bool(r[d]):
                off_rows.append({'name': r['name'], 'day': d, 'type': 'OFF', 'shift': 'ANY', 'weight': 999})

    combined = pd.concat([req_df_in, pd.DataFrame(off_rows)], ignore_index=True)
    combined.to_csv(tmp_path, index=False)


# Need these at runtime even if expander not opened
avail_for_build = pd.read_csv(AVAIL_PATH)
for d in DAYS:
    if d in avail_for_build.columns:
        avail_for_build[d] = avail_for_build[d].fillna(1).astype(int).astype(bool)
    else:
        avail_for_build[d] = True
avail_for_build['name'] = avail_for_build['name'].astype(str).str.strip()

if REQ_PATH.exists():
    req_for_build = pd.read_csv(REQ_PATH)
else:
    req_for_build = pd.DataFrame(columns=['name', 'day', 'type', 'shift', 'weight'])

if generate or reroll:
    if reroll:
        st.session_state.seed = int(st.session_state.seed) + 1

    # Reload model inputs (in case advanced settings changed)
    staff_df = ensure_staff_columns(pd.read_csv(STAFF_PATH))
    staff_df.to_csv(STAFF_PATH, index=False)

    people = load_staff(STAFF_PATH)
    rules = load_rules(RULES_PATH)
    week = load_week_template(WEEK_PATH)

    tmp_req = BASE / '_requests_runtime.csv'
    _build_combined_requests(tmp_req, avail_for_build, req_for_build)
    requests = load_requests(tmp_req)

    try:
        assignments = solve_week(
            people=people,
            week=week,
            rules=rules,
            requests=requests,
            random_seed=int(st.session_state.seed),
        )
        st.session_state.roster = assignments
        st.success('Generated roster (best effort).')
    except Exception as e:
        st.session_state.roster = None
        st.error(f'Generate failed: {e}')


# ============================================================
# Preview + Manual hole filling (no recursion)
# ============================================================
if st.session_state.roster is not None:
    staff_df = ensure_staff_columns(pd.read_csv(STAFF_PATH))
    week = load_week_template(WEEK_PATH)
    roster = st.session_state.roster

    holes = [(d, s) for d in DAYS for s in build_shift_order(week) if s in week.get(d, []) and roster.get((d, s)) is None]

    st.subheader('Roster preview')
    st.caption('Auto-generated roster. Holes show as —. Use Repair view to fill holes (no solver rerun).')

    view = st.radio('View', ['Published view (like Excel)', 'Repair view (fill holes)'], horizontal=True)

    if view.startswith('Published'):
        g_df = published_table_for_sheet(staff_df, roster, week, 'show_gondola', GONDOLA_SHIFTS)
        s_df = published_table_for_sheet(staff_df, roster, week, 'show_gs', GS_SHIFTS)

        st.markdown('### Gondola')
        st.dataframe(g_df, use_container_width=True, hide_index=True)

        st.markdown('### Guest Services')
        st.dataframe(s_df, use_container_width=True, hide_index=True)

    else:
        table = assignments_to_table(roster, week)
        st.dataframe(style_holes(table), use_container_width=True)

        st.write(f'Holes: **{len(holes)}**')

        if holes:
            with st.expander('Fill a hole (default suggests skilled staff)', expanded=True):
                people_names = [p.name for p in load_staff(STAFF_PATH)]

                hole_days = sorted({d for d, _ in holes}, key=lambda x: DAYS.index(x))
                day = st.selectbox('Day', hole_days, key='hole_day')

                hole_shifts = [s for (d, s) in holes if d == day]
                shift = st.selectbox('Shift', hole_shifts, key='hole_shift')

                show_all = st.checkbox('Show all staff (override skills)', value=False)

                options = people_not_working_that_day(roster, day, people_names)

                if not show_all:
                    if shift in GONDOLA_SHIFTS:
                        can_set = set(staff_df.loc[staff_df['can_gondola'] == 1, 'name'].tolist())
                        options = [p for p in options if p in can_set]
                    elif shift in GS_SHIFTS:
                        can_set = set(staff_df.loc[staff_df['can_gs'] == 1, 'name'].tolist())
                        options = [p for p in options if p in can_set]

                options = [''] + options
                choice = st.selectbox('Assign staff', options, key='hole_assign')

                if choice:
                    roster[(day, shift)] = choice
                    st.session_state.roster = roster
                    st.success(f'Assigned {choice} to {day} {shift}')
                    st.rerun()
        else:
            st.success('No holes 🎉')

    # ============================================================
    # Export
    # ============================================================
    st.subheader('Export')
    if st.button('⬇️ Export Excel (uses current preview, including overrides)'):
        runtime_template = make_runtime_template()
        gondola_sheet, gs_sheet = detect_sheets(runtime_template)
        apply_visibility_to_runtime_template(runtime_template, gondola_sheet, gs_sheet, staff_df)

        out = output_filename()
        export_roster_to_template(
            assignments=st.session_state.roster,
            template_path=runtime_template,
            output_path=out,
            gondola_sheet=gondola_sheet,
            gs_sheet=gs_sheet,
            gondola_gs_label='GS Host',
        )

        with open(out, 'rb') as f:
            st.download_button('Download roster', f, file_name=out.name)

else:
    st.info('Press **Generate** to create a roster preview.')
